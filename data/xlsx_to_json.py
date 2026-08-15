"""
DBACS – Excel → JSON Konvertierung
===================================
Exportiert folgende Sheets aus ga_komponenten.xlsx:
  - kabel_nym_j           → kabel_nym_j.json           (nur aktive Einträge)
  - wandschraenke         → wandschraenke.json          (nur aktive Einträge)
  - kabelzugschellen      → kabelzugschellen.json       (nur aktive Einträge)
  - standschraenke        → standschraenke.json         (nur aktive Einträge)
  - sockel                → sockel.json                 (nur aktive Einträge)
  - bodenbleche           → bodenbleche.json            (nur aktive Einträge)
  - einzelbauteile        → einzelbauteile.json         (nur aktive Einträge)
  - baugruppen +
    baugruppen_bauteile   → baugruppen.json             (nur aktive Baugruppen;
                                                          bauteile[] aus der
                                                          Verknüpfungstabelle
                                                          zusammengebaut)
  - feldgeraete           → feldgeraete.json             (nur aktive Einträge;
                                                          externe Betriebsmittel
                                                          außerhalb des Schalt-
                                                          schranks, von Modul 5
                                                          genutzt, Session 51
                                                          Nachtrag 7)

Reine Referenz-Sheets (kein Export, keine eigene JSON-Datei):
  - funktionsbereiche     Klartext-Nachschlagewerk zu 'gewerk'/'funktionsbereich'
                          (baugruppen) – 11 DIN-276-Kategorien.
  - zonen                 Klartext-Nachschlagewerk zu 'zone' (einzelbauteile,
                          baugruppen_bauteile) – physische Platzierungszonen.
                          Beide Sheets sind rein für die menschliche Lesbarkeit
                          in Excel gedacht; die Variablen/Codes selbst bleiben
                          die Datengrundlage für xlsx_to_json.py und die Module.

'einzelbauteile.zone' (Session 44): immer als Array exportiert, auch bei nur
einer Zone – Excel-Zelle ist eine Komma-Liste (z.B. "klemm_l,klemm_f,klemm_s"
bei den PT-Klemmen, die baugleich in mehreren Abgangs-Klemmenzonen verwendbar
sind, ohne die Katalogzeile zu verdreifachen). Erster Eintrag ist der Default.
Modul 4 laesst den Nutzer bei Direktbauteil-Auswahl explizit waehlen, welche
der erlaubten Zonen fuer die jeweilige Menge gilt (#zone_auswahl), analog zum
laenger bestehenden bt.zone-Override bei Baugruppen-Bauteilen (der bleibt ein
einzelner String je Verwendung, kein Array). Siehe CLAUDE.md Session 44.
  - planungsfabrikate     Kategorie -> bevorzugter Hersteller (Rittal/Siemens/
                          Phoenix Contact/Dehn), reine Dokumentation.

'reiheneinbaugeraete' (Sheet + reiheneinbaugeraete.json) in Session 40 komplett
entfernt: unverifizierte Session-20-Altlast, nie von einem Modul geladen,
enthielt Dubletten zu bereits besser modellierten einzelbauteile-Eintraegen.
Siehe CLAUDE.md Session 40.

Namenskonvention Excel-Spalten vs. JSON-Ausgabe (Session 39): 'artikel_nr'
(Excel-Spalte, einheitlich für alle Artikel-/Bestellnummern-Felder über alle
Sheets, ersetzt das frühere uneinheitliche 'bestellnummer') und
'preis_stueck_eur' (Excel-Spalte, ersetzt 'preis_stueckpreis_eur') werden beim
Einlesen zwar konsistent benannt, die bereits bestehenden JSON-Ausgabeschlüssel
('bestellnummer', 'preis_stueckpreis_eur') bleiben bewusst unverändert, um
keine Aenderungen an den Modul-1-4/7-JS-Dateien zu erzwingen, die diese Keys
bereits konsumieren.

Aufruf (aus data/-Verzeichnis):
    python3 xlsx_to_json.py

Abhängigkeit:
    pip install openpyxl
"""

import json
import math
import openpyxl
from pathlib import Path

EXCEL_FILE = Path(__file__).parent / 'ga_komponenten.xlsx'

# DDC-Datenpunkt-Felder je Bauteil (nur bei automationsanbindung=true befüllt):
# bei ddc_io-Modulen = bereitgestellte Kapazität je Typ, sonst = Bedarf je Typ.
# dp_fb_* (kommunikative/Feldbus-Datenpunkte) brauchen zusaetzlich das Feld
# 'feldbus_protokoll' (Session 41): welches Protokoll dieser Datenpunkt-
# Bedarf/diese -Kapazitaet nutzt - gueltige Werte EXAKT 'mbus', 'modbus_rtu'
# oder 'modbus_tcp' (muessen woertlich zu Modul 4s FB_SUMMARY_GROUPS-Keys
# passen). Ein Modbus-RTU-Bedarf darf nur durch Modbus-RTU-Kapazitaet gedeckt
# werden, nicht durch M-Bus o.ae. - Modul 4 fuehrt dafuer 3 getrennte
# Kapazitaets-/Bedarfs-Gruppen (Eingabeleiste "DDC-Automationseinrichtung").
# 'dp_beschreibung' (Freitext, Session 41): erklaert in Klartext, welche
# konkreten Datenpunkte sich hinter den dp_*/dp_fb_*-Zahlen verbergen (z.B.
# "physikalisch: BI=Rueckmeldung Betrieb, BI=Rueckmeldung Stoerung, DO=Schalt-
# befehl; kommunikativ (Modbus RTU): AI=Wirkleistung, AI=Energiezaehler").
# 'keine_platzierung_mp' (Boolean, Session 41): Bauteil wird auf ein anderes
# Bauteil aufgesteckt (z.B. Hilfsschalterblock auf ein Schuetz) und braucht
# dadurch KEIN eigenes TE-Feld auf der Montageplatte - erscheint in Modul 4
# weiterhin in der Stueckliste, wird aber nicht im Schrankbild platziert/
# gezeichnet.
# 'zubehoer_artikel_nr' (Text, Session 41 Nachtrag 5): Grundausstattung -
# artikel_nr eines Bauteils, das automatisch in gleicher Menge mitgezaehlt
# wird, sobald DIESES Bauteil in Modul 4 hinzugefuegt wird (z.B. Schuetz ->
# Hilfsschalterblock). Kein Stromlaufplan-Anspruch, nur Platzbedarf/Preis -
# Nutzer-Vorgabe.
DP_FELDER = ['dp_ai', 'dp_ao', 'dp_bi', 'dp_bo', 'dp_fb_ai', 'dp_fb_ao', 'dp_fb_bi', 'dp_fb_bo']


def export_kabel_nym_j(wb):
    SHEET = 'kabel_nym_j'
    JSON_FILE = Path(__file__).parent / 'kabel_nym_j.json'

    if SHEET not in wb.sheetnames:
        print(f'FEHLER: Sheet "{SHEET}" nicht in der Excel-Datei.')
        return

    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get('aktiv'):
            continue
        rows.append({
            'typ':             str(rec['typ']),
            'n_adern':         int(rec['n_adern']),
            'querschnitt_mm2': float(rec['querschnitt_mm2']),
            'd_aussen_mm':     float(rec['d_aussen_mm']),
            'bezeichnung':     str(rec['bezeichnung']),
        })

    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f'{len(rows)} Kabeleintraege exportiert → {JSON_FILE.name}')


def export_wandschraenke(wb):
    SHEET = 'wandschraenke'
    JSON_FILE = Path(__file__).parent / 'wandschraenke.json'

    if SHEET not in wb.sheetnames:
        print(f'HINWEIS: Sheet "{SHEET}" nicht gefunden – uebersprungen.')
        return

    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get('aktiv'):
            continue
        rows.append({
            'hersteller':            str(rec['hersteller']),
            'bezeichnung':           str(rec['bezeichnung']),
            'bestellnummer':         str(rec['artikel_nr']),
            'b_gehaeuse_aussen_mm':  int(rec['b_gehaeuse_aussen_mm']),
            'h_gehaeuse_aussen_mm':  int(rec['h_gehaeuse_aussen_mm']),
            't_gehaeuse_aussen_mm':  int(rec['t_gehaeuse_aussen_mm']),
            'b_mplatte_mm':          int(rec['b_mplatte_mm']),
            'h_mplatte_mm':          int(rec['h_mplatte_mm']),
            'preis_stueckpreis_eur': float(rec['preis_stueck_eur']) if rec.get('preis_stueck_eur') is not None else None,
            'preis_lieferung_eur':   float(rec['preis_lieferung_eur'])   if rec.get('preis_lieferung_eur')   is not None else None,
            'preis_montage_eur':     float(rec['preis_montage_eur'])     if rec.get('preis_montage_eur')     is not None else None,
            'preis_gesamt_eur':      float(rec['preis_gesamt_eur'])      if rec.get('preis_gesamt_eur')      is not None else None,
        })

    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f'{len(rows)} Wandschraenke exportiert → {JSON_FILE.name}')


def export_kabelzugschellen(wb):
    SHEET = 'kabelzugschellen'
    JSON_FILE = Path(__file__).parent / 'kabelzugschellen.json'

    if SHEET not in wb.sheetnames:
        print(f'HINWEIS: Sheet "{SHEET}" nicht gefunden – uebersprungen.')
        return

    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get('aktiv'):
            continue
        rows.append({
            'hersteller':          str(rec['hersteller']),
            'bezeichnung':         str(rec['bezeichnung']),
            'bestellnummer':       str(rec['artikel_nr']),
            'd_kabel_min_mm':      float(rec['d_kabel_min_mm']),
            'd_kabel_max_mm':      float(rec['d_kabel_max_mm']),
            'h_schelle_mm':        float(rec['h_schelle_mm']),
            'b_schelle_mm':        float(rec['b_schelle_mm']),
            't_schelle_mm':        float(rec['t_schelle_mm']),
            'preis_stueckpreis_eur': float(rec['preis_stueck_eur']) if rec.get('preis_stueck_eur') is not None else None,
            'preis_lieferung_eur':   float(rec['preis_lieferung_eur'])   if rec.get('preis_lieferung_eur')   is not None else None,
            'preis_montage_eur':     float(rec['preis_montage_eur'])     if rec.get('preis_montage_eur')     is not None else None,
            'preis_gesamt_eur':      float(rec['preis_gesamt_eur'])      if rec.get('preis_gesamt_eur')      is not None else None,
        })

    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f'{len(rows)} Kabelzugschellen exportiert → {JSON_FILE.name}')


def export_standschraenke(wb):
    SHEET = 'standschraenke'
    JSON_FILE = Path(__file__).parent / 'standschraenke.json'

    if SHEET not in wb.sheetnames:
        print(f'HINWEIS: Sheet "{SHEET}" nicht gefunden – uebersprungen.')
        return

    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get('aktiv'):
            continue
        rows.append({
            'hersteller':            str(rec['hersteller']),
            'bezeichnung':           str(rec['bezeichnung']),
            'bestellnummer':         str(rec['artikel_nr']),
            'b_gehaeuse_aussen_mm':  int(rec['b_gehaeuse_aussen_mm']),
            'h_gehaeuse_aussen_mm':  int(rec['h_gehaeuse_aussen_mm']),
            't_gehaeuse_aussen_mm':  int(rec['t_gehaeuse_aussen_mm']),
            'b_mplatte_mm':          int(rec['b_mplatte_mm']),
            'h_mplatte_mm':          int(rec['h_mplatte_mm']),
            'preis_stueckpreis_eur': float(rec['preis_stueck_eur']) if rec.get('preis_stueck_eur') is not None else None,
            'preis_lieferung_eur':   float(rec['preis_lieferung_eur'])   if rec.get('preis_lieferung_eur')   is not None else None,
            'preis_montage_eur':     float(rec['preis_montage_eur'])     if rec.get('preis_montage_eur')     is not None else None,
            'preis_gesamt_eur':      float(rec['preis_gesamt_eur'])      if rec.get('preis_gesamt_eur')      is not None else None,
        })

    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f'{len(rows)} Standschraenke exportiert → {JSON_FILE.name}')


def export_sockel(wb):
    SHEET = 'sockel'
    JSON_FILE = Path(__file__).parent / 'sockel.json'

    if SHEET not in wb.sheetnames:
        print(f'HINWEIS: Sheet "{SHEET}" nicht gefunden – uebersprungen.')
        return

    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get('aktiv'):
            continue
        rows.append({
            'hersteller':            str(rec['hersteller']),
            'bezeichnung':           str(rec['bezeichnung']),
            'bestellnummer':         str(rec['artikel_nr']),
            'b_gehaeuse_aussen_mm':  int(rec['b_gehaeuse_aussen_mm']),
            'h_sockel_mm':           int(rec['h_sockel_mm']),
            'preis_stueckpreis_eur': float(rec['preis_stueck_eur']) if rec.get('preis_stueck_eur') is not None else None,
            'preis_lieferung_eur':   float(rec['preis_lieferung_eur'])   if rec.get('preis_lieferung_eur')   is not None else None,
            'preis_montage_eur':     float(rec['preis_montage_eur'])     if rec.get('preis_montage_eur')     is not None else None,
            'preis_gesamt_eur':      float(rec['preis_gesamt_eur'])      if rec.get('preis_gesamt_eur')      is not None else None,
        })

    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f'{len(rows)} Sockel exportiert → {JSON_FILE.name}')


def export_bodenbleche(wb):
    SHEET = 'bodenbleche'
    JSON_FILE = Path(__file__).parent / 'bodenbleche.json'

    if SHEET not in wb.sheetnames:
        print(f'HINWEIS: Sheet "{SHEET}" nicht gefunden – uebersprungen.')
        return

    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get('aktiv'):
            continue
        rows.append({
            'hersteller':            str(rec['hersteller']),
            'bezeichnung':           str(rec['bezeichnung']),
            'bestellnummer':         str(rec['artikel_nr']),
            'b_gehaeuse_aussen_mm':  int(rec['b_gehaeuse_aussen_mm']),
            'anzahl_platten':        int(rec['anzahl_platten']),
            'preis_stueckpreis_eur': float(rec['preis_stueck_eur']) if rec.get('preis_stueck_eur') is not None else None,
            'preis_lieferung_eur':   float(rec['preis_lieferung_eur'])   if rec.get('preis_lieferung_eur')   is not None else None,
            'preis_montage_eur':     float(rec['preis_montage_eur'])     if rec.get('preis_montage_eur')     is not None else None,
            'preis_gesamt_eur':      float(rec['preis_gesamt_eur'])      if rec.get('preis_gesamt_eur')      is not None else None,
        })

    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f'{len(rows)} Bodenblech-Saetze exportiert → {JSON_FILE.name}')


def export_einzelbauteile(wb):
    SHEET = 'einzelbauteile'
    JSON_FILE = Path(__file__).parent / 'einzelbauteile.json'

    if SHEET not in wb.sheetnames:
        print(f'HINWEIS: Sheet "{SHEET}" nicht gefunden – uebersprungen.')
        return

    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get('aktiv'):
            continue
        b_mm = float(rec['b_mm']) if rec.get('b_mm') is not None else None
        entry = {
            'artikel_nr':  str(rec['artikel_nr']),
            'bezeichnung': str(rec['bezeichnung']),
            'hersteller':  str(rec['hersteller']),
            'bauteil_typ': str(rec['bauteil_typ']),
            'b_mm':        b_mm,
            'te_breite':   math.ceil(b_mm / 18) if b_mm is not None else None,
            'h_mm':        float(rec['h_mm']) if rec.get('h_mm') is not None else None,
            # Session 44: zone ist immer ein Array erlaubter Zonen, erster Eintrag
            # ist der Default (z.B. PT-Klemmen: klemm_l,klemm_f,klemm_s) - auch bei
            # nur einer Zone, damit Modul 4 nicht zwischen Skalar/Array unterscheiden muss.
            'zone':        [z.strip() for z in str(rec['zone']).split(',')] if rec.get('zone') is not None else None,
        }
        if rec.get('kategorie') is not None:
            entry['kategorie'] = str(rec['kategorie'])
        if rec.get('einbaulage') is not None:
            entry['einbaulage'] = str(rec['einbaulage'])
        if rec.get('automationsanbindung'):
            entry['automationsanbindung'] = True
            for dp_feld in DP_FELDER:
                if rec.get(dp_feld):
                    entry[dp_feld] = int(rec[dp_feld])
            if rec.get('feldbus_protokoll') is not None:
                entry['feldbus_protokoll'] = str(rec['feldbus_protokoll'])
        if rec.get('dp_beschreibung') is not None:
            entry['dp_beschreibung'] = str(rec['dp_beschreibung'])
        if rec.get('keine_platzierung_mp'):
            entry['keine_platzierung_mp'] = True
        if rec.get('zubehoer_artikel_nr') is not None:
            entry['zubehoer_artikel_nr'] = str(rec['zubehoer_artikel_nr'])
        if rec.get('lvb_integriert') is not None:
            entry['lvb_integriert'] = bool(rec['lvb_integriert'])
        # doppelstock_variante_artikel_nr (Session 51): nur bei Standard-
        # Klemmen mit Doppelstock-Gegenstueck gesetzt - resolveBaugruppenBauteile()
        # in Modul 4 ersetzt bei aktiver Doppelstock-Option ein Klemmenpaar
        # (Signal+Referenz derselben DDC-Reserve-Baugruppe) durch EINE
        # Doppelstockklemme im selben mm-Platzbedarf.
        if rec.get('doppelstock_variante_artikel_nr') is not None:
            entry['doppelstock_variante_artikel_nr'] = str(rec['doppelstock_variante_artikel_nr'])
        # trennklemme_variante_artikel_nr (Session 51 Nachtrag): verweist von
        # der Standardklemme auf ihre Trennklemmen-Variante (analog zu
        # doppelstock_variante_artikel_nr) - zusammen ergeben beide Felder die
        # 4 waehlbaren Klemmentypen (Standard/Doppelstock/Trennklemme/
        # Doppelstock-Trennklemme) in resolveKlemmeArtikel() (Modul 4).
        if rec.get('trennklemme_variante_artikel_nr') is not None:
            entry['trennklemme_variante_artikel_nr'] = str(rec['trennklemme_variante_artikel_nr'])
        # max_ea_module/ddc_netzteil_artikel_nr/ddc_sicherung_artikel_nr (Session 50):
        # nur bei ddc_cpu-Bauteilen relevant. Herstellerseitige Obergrenze der
        # E/A-Module je Automationsstation (max_ea_module) sowie die fest
        # zugehoerige Versorgung (Netzteil+Vorsicherung) - jede CPU bringt beim
        # automatischen Ergaenzen (buildQueues() in Modul 4) ihre eigene mit,
        # statt sich eine gemeinsame mit anderen CPUs zu teilen.
        if rec.get('max_ea_module') is not None:
            entry['max_ea_module'] = int(rec['max_ea_module'])
        if rec.get('ddc_netzteil_artikel_nr') is not None:
            entry['ddc_netzteil_artikel_nr'] = str(rec['ddc_netzteil_artikel_nr'])
        if rec.get('ddc_sicherung_artikel_nr') is not None:
            entry['ddc_sicherung_artikel_nr'] = str(rec['ddc_sicherung_artikel_nr'])
        # auto_ea_cpu (Session 50): seit der Ergaenzung mehrerer ddc_cpu-Typen
        # (Kompakt PXC4.E16.A/PXC5.E24.A + Modular PXC7.E400.A) muss buildQueues()
        # in Modul 4 eindeutig wissen, welche CPU sie fuer die automatische
        # Netzteil->CPU->E/A-Modul-Ergaenzung verwendet - ohne dieses Flag waere
        # die Wahl von der zufaelligen Katalog-Reihenfolge abhaengig. Nur bei
        # PXC7.E400.A gesetzt (0 Onboard-E/A, passt zum DBACS-Prinzip "Verbrauch
        # entsteht ausschliesslich an TXM-Modulen").
        if rec.get('auto_ea_cpu'):
            entry['auto_ea_cpu'] = True
        if rec.get('klemmen_zusatz') is not None:
            entry['klemmen_zusatz'] = int(rec['klemmen_zusatz'])
        if rec.get('preis_stueck_eur') is not None:
            entry['preis_eur'] = float(rec['preis_stueck_eur'])
        if rec.get('preis_lieferung_eur') is not None:
            entry['preis_lieferung_eur'] = float(rec['preis_lieferung_eur'])
        if rec.get('montage_minuten') is not None:
            entry['montage_minuten'] = float(rec['montage_minuten'])
        if rec.get('preis_gesamt_eur') is not None:
            entry['preis_gesamt_eur'] = float(rec['preis_gesamt_eur'])
        entry['geprueft'] = bool(rec.get('geprueft'))
        rows.append(entry)

    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f'{len(rows)} Einzelbauteile exportiert → {JSON_FILE.name}')


def export_baugruppen(wb):
    # betriebsmittel (Session 49, Freitext, optional): externes Feldgeraet/
    # Betriebsmittel, das ueber die Abgangs-/Eingangsklemmen dieser Baugruppe
    # angeschlossen wird (z.B. "Pumpe") -- kein Bestandteil des Schaltschranks
    # selbst, dient spaeter als Grundlage fuer eine separate Betriebsmittel-
    # Stueckliste. Bewusst Freitext, keine feste Kategorie-Liste.
    # Feld-Korrektur Session 37: 'funktionsbereiche' (Komma-Liste) -> 'funktionsbereich'
    # (einzelner Klartext-Wert, Companion zu 'gewerk'); 'automationsfunktionen'
    # (Komma-Liste) -> 'automationsanbindung' (Boolean, gleicher Name/gleiche
    # Semantik wie bei einzelbauteile). 'gewerk' selbst ist unveraendert als
    # Spalte, soll aber laut Nutzer-Entscheidung kuenftig den 3-stelligen
    # DIN-276-Code als Text tragen (z.B. "430") statt eines Kurznamens wie
    # "lueftung" -- diese inhaltliche Migration der Bestandsdaten ist bewusst
    # noch NICHT Teil dieser Aenderung (erst Feld-Struktur, dann Inhalte),
    # siehe CLAUDE.md.
    SHEET = 'baugruppen'
    JOIN_SHEET = 'baugruppen_bauteile'
    JSON_FILE = Path(__file__).parent / 'baugruppen.json'

    if SHEET not in wb.sheetnames:
        print(f'HINWEIS: Sheet "{SHEET}" nicht gefunden – uebersprungen.')
        return
    if JOIN_SHEET not in wb.sheetnames:
        print(f'HINWEIS: Sheet "{JOIN_SHEET}" nicht gefunden – uebersprungen.')
        return

    # Verknüpfungstabelle einmal einlesen und nach bg_id gruppieren
    join_ws = wb[JOIN_SHEET]
    join_headers = [cell.value for cell in next(join_ws.iter_rows(min_row=1, max_row=1))]
    bauteile_je_bg = {}
    for row in join_ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(join_headers, row))
        bg_id = rec.get('bg_id')
        if not bg_id or not rec.get('artikel_nr'):
            continue
        bt = {'artikel_nr': str(rec['artikel_nr']), 'menge': int(rec['menge'])}
        if rec.get('zone') is not None:
            bt['zone'] = str(rec['zone'])
        if rec.get('zeilenumbruch_davor'):
            bt['rowBreak'] = True
        # neue_gruppe (Session 50): markiert den Beginn einer Automations-
        # Gruppe (Netzteil -> CPU -> E/A-Module) - erzwingt IMMER eine frische
        # Hutschienenreihe, auch wenn die zuletzt per rowBreak erzwungene Reihe
        # noch Platz haette (anders als rowBreak allein, das an eine
        # bestehende erzwungene Reihe anhaengt, siehe assignDevicesToRows()
        # in Modul 4).
        if rec.get('neue_gruppe'):
            bt['groupStart'] = True
        # dp_ai/ao/bi/bo (Session 49): DDC-Datenpunktbedarf je Verwendung
        # dieses Bauteils in DIESER Baugruppe, analog zum zone-Override –
        # z.B. eine Klemme traegt selbst keinen DDC-Bezug, stellt aber in
        # einer "Reserve"-Baugruppe einen vorbereiteten Anschlusspunkt dar.
        for dpf in ('dp_ai', 'dp_ao', 'dp_bi', 'dp_bo'):
            if rec.get(dpf):
                bt[dpf] = int(rec[dpf])
        # lvb_erforderlich (Session 49): dieser Datenpunktbedarf darf nur durch
        # ein Modul MIT integrierter lokaler Vorrangbedienung gedeckt werden.
        if rec.get('lvb_erforderlich'):
            bt['lvb_erforderlich'] = True
        bauteile_je_bg.setdefault(str(bg_id), []).append(bt)

    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get('aktiv'):
            continue
        entry = {
            'id':          str(rec['id']),
            'name':        str(rec['name']),
            'gewerk':      str(rec['gewerk']),
            'beschreibung': str(rec['beschreibung']),
            'bauteile':    bauteile_je_bg.get(str(rec['id']), []),
        }
        # funktionsbereich ist seit Session 51 Nachtrag 9 immer ein Array
        # (Komma-Liste in Excel), auch bei nur einem Wert - analog zu
        # einzelbauteile.zone (Session 44): eine Baugruppe kann in mehreren
        # Gewerke-Tabs gleichzeitig auswaehlbar sein (z.B. ein Raumsensor in
        # Heizung/Lueftung/Kaelte), ohne die Katalogzeile zu vervielfachen.
        if rec.get('funktionsbereich') is not None:
            entry['funktionsbereich'] = [f.strip() for f in str(rec['funktionsbereich']).split(',')]
        if rec.get('betriebsmittel'):
            entry['betriebsmittel'] = str(rec['betriebsmittel'])
        if rec.get('feldgeraet_artikel_nr'):
            entry['feldgeraet_artikel_nr'] = str(rec['feldgeraet_artikel_nr'])
        if rec.get('automationsanbindung'):
            entry['automationsanbindung'] = True
        entry['geprueft'] = bool(rec.get('geprueft'))
        rows.append(entry)

    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f'{len(rows)} Baugruppen exportiert → {JSON_FILE.name}')


def export_feldgeraete(wb):
    # Feldgeraete-Katalog (Session 51 Nachtrag 7): externe Betriebsmittel
    # (Pumpen, Feldsensoren/-aktoren) AUSSERHALB des Schaltschranks - analog
    # 'einzelbauteile', aber bewusst OHNE b_mm/h_mm/te_breite/zone (werden
    # nicht in der Schaltschrank-SVG platziert). Von 'baugruppen.
    # feldgeraet_artikel_nr' referenziert, gespeist in Modul 5 zu einer
    # eigenen, bepreisten Feldgeraete-Stueckliste.
    SHEET = 'feldgeraete'
    JSON_FILE = Path(__file__).parent / 'feldgeraete.json'

    if SHEET not in wb.sheetnames:
        print(f'HINWEIS: Sheet "{SHEET}" nicht gefunden – uebersprungen.')
        return

    ws = wb[SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        if not rec.get('aktiv'):
            continue
        entry = {
            'artikel_nr':  str(rec['artikel_nr']),
            'bezeichnung': str(rec['bezeichnung']),
            'hersteller':  str(rec['hersteller']),
            'kategorie':   str(rec['kategorie']),
        }
        if rec.get('preis_stueck_eur') is not None:
            entry['preis_eur'] = float(rec['preis_stueck_eur'])
        if rec.get('quelle_hinweis'):
            entry['quelle_hinweis'] = str(rec['quelle_hinweis'])
        # zubehoer_feldgeraet_artikel_nr (Lueftungssensoren-Session): Pflichtzubehoer,
        # das NICHT als eigene Baugruppe waehlbar ist (z.B. Montagekonsole eines
        # Kanalrauchmelders), aber automatisch mit der Menge des Hauptgeraets in die
        # Feldgeraete-Stueckliste (Modul 5) einfliesst - analog zu einzelbauteile.
        # zubehoer_artikel_nr/syncZubehoer() in Modul 4, nur auf Feldgeraete-Ebene
        # statt Schaltschrank-Ebene. zubehoer_menge = Multiplikator je Hauptgeraet
        # (Default 1, falls nicht gesetzt).
        if rec.get('zubehoer_feldgeraet_artikel_nr'):
            entry['zubehoer_feldgeraet_artikel_nr'] = str(rec['zubehoer_feldgeraet_artikel_nr'])
            entry['zubehoer_menge'] = int(rec['zubehoer_menge']) if rec.get('zubehoer_menge') is not None else 1
        entry['geprueft'] = bool(rec.get('geprueft'))
        rows.append(entry)

    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f'{len(rows)} Feldgeraete exportiert → {JSON_FILE.name}')


def main():
    if not EXCEL_FILE.exists():
        print(f'FEHLER: {EXCEL_FILE} nicht gefunden.')
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    export_kabel_nym_j(wb)
    export_wandschraenke(wb)
    export_kabelzugschellen(wb)
    export_standschraenke(wb)
    export_sockel(wb)
    export_bodenbleche(wb)
    export_einzelbauteile(wb)
    export_baugruppen(wb)
    export_feldgeraete(wb)


if __name__ == '__main__':
    main()
